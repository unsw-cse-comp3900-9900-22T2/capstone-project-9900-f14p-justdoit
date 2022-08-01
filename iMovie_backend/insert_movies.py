import pymysql
import xlrd
import datetime
from openpyxl.reader.excel import load_workbook
from builtins import int

import random
import time
# conn = pymysql.connect(host='localhost',user='root',passwd='qwer1234',db="imovie")
conn = pymysql.connect(host='rm-bp16q3qua05mfx407lo.mysql.rds.aliyuncs.com',user='root',passwd='1234@qwer!',db="imovie")
# sql = 'select * from movies'
# conn.ping(reconnect=True)  # reconnecting mysql
cur = conn.cursor()
# cur.execute(sql)
# conn.commit()





def randomString(num):
    a = random.sample('ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789', num)
    return ''.join(a)
def getTime():

    time_stamp = time.localtime(time.time()) # 132131232434
    time_form = time.strftime('%Y-%m-%d %H:%M:%S', time_stamp)  # 2022-6-27 12:13:00

    time_stamp = int(time.mktime(time_stamp))
    return [time_form, time_stamp]
def getUniqueid():
    time_stamp = getTime()[1]
    uniqueId = str(randomString(10)) + str(time_stamp)
    return uniqueId

workbook = load_workbook("./movies_data.xlsx")
# sheets = workbook.get_sheet_names()
# worksheet = workbook.get_sheet_by_name(sheets[0])
sheets = workbook.sheetnames
worksheet = workbook["Sheet1"]
for row in worksheet.rows:
    sqlstr = []
    for cell in row:
        sqlstr.append(cell.value)
    mid = getUniqueid()
    data_time = getTime()[0]
    if sqlstr[8]:
        sqlstr_8 = sqlstr[8].lower()
    else:
        sqlstr_8 = None
    # print(str(sqlstr[1]), sqlstr[5])

    valuestr = [str(mid), str(sqlstr[1]),int(sqlstr[2]), str(sqlstr[3]), str(sqlstr[4]), int(sqlstr[5])
                , str(sqlstr[6]), str(sqlstr[7]), str(sqlstr_8),str(sqlstr[9]),str(sqlstr[10]),1,data_time,data_time,sqlstr[11]]

    cur.execute("insert into movies(mid,moviename,year,director,description,duration,country,language,genre,cast,coverimage,active,ctime,utime,release_date)value(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",valuestr)



import hashlib


#
def EnPassWord(password: str) -> str:
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

admin_uid = getUniqueid()
admin_name = "admin"
admin_email = "adminemail@mail.com"
admin_password = "admin1234"
en_password = EnPassWord(admin_password)
admin_role = 1
admin_date = getTime()[0]
admin_ctime = admin_date
admin_utime = admin_date

admin_valuestr = [str(admin_uid) ,str(admin_name), str(en_password),admin_email, int(admin_role), 1,admin_ctime,admin_utime]

cur.execute("insert into users(uid,username,password,email,role,active,ctime,utime)value(%s,%s,%s,%s,%s,%s,%s,%s)",admin_valuestr)

# admin = []

cur.close()
conn.commit()
conn.close()
